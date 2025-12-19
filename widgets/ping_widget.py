import threading

from PySide6.QtCore import QRegularExpression, Qt
from PySide6.QtGui import QRegularExpressionValidator, QColor
from PySide6.QtWidgets import QWidget, QGridLayout, QLabel
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from qfluentwidgets import InfoBar, InfoBarIcon, InfoBarPosition, ToolTipFilter, ToolTipPosition
from qframelesswindow.utils import getSystemAccentColor

from design.Ui_PingWidget import Ui_PingWidget
from widgets.ping_worker import PingWorker


class PingWidget(QWidget):
    def __init__(self, parent):
        super(PingWidget, self).__init__()
        self.ui = Ui_PingWidget()
        self.ui.setupUi(self)
        self.parent = parent
        self.ip_prefix = ""
        self.ping_worker = None
        self.ping_finished = False
        self.wb = Workbook()
        self.ws = None
        self.initialExcel()

        self.progressBarStyleSheet = f"""
            QProgressBar {{
                border-radius: 4px;
                text-align: center;
                }}
            QProgressBar::chunk {{
                background: { getSystemAccentColor().name() };
                border-radius: 4px;
                border: 1px solid rgba(255, 255, 255, 30%);
                box-shadow: inset 0 1px 2px rgba(0, 0, 0, 0.1);
                }}"""
        self.ui.progressBar.setStyleSheet(self.progressBarStyleSheet)
        self.result_labels = []
        self.initialResultGrid()

        # 设置IP地址输入验证器
        regex = QRegularExpression("^(0|[1-9]\\d?|1\\d{2}|2[0-4]\\d|25[0-5])$")
        validator = QRegularExpressionValidator(regex)
        self.ui.ipEdit_1.setValidator(validator)
        self.ui.ipEdit_2.setValidator(validator)
        self.ui.ipEdit_3.setValidator(validator)

        # 连接按钮事件
        self.ui.startPingBtn.clicked.connect(self.startPingTest)
        self.ui.stopPingBtn.clicked.connect(self.stopPingTest)

    def initialExcel(self):
        self.ws = self.wb.active
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 8
        self.ws.column_dimensions['C'].width = 12
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 15
        self.ws.column_dimensions['F'].width = 15
        headers = ["IP", "状态", "丢包率%", "最大延迟ms", "最小延迟ms", "平均延迟ms"]
        for col, header in enumerate(headers, 1):
            self.cell_format(self.ws.cell(row=1, column=col, value=header))

    def initialResultGrid(self):
        """填充结果区数据"""
        grid = QGridLayout(self.ui.tableWidget)
        grid.setSpacing(4)
        grid.setContentsMargins(0, 0, 0, 0)

        for row in range(16):
            col_list = []
            for col in range(16):
                index = row * 16 + col
                label = ResultLabel(str(index))
                grid.addWidget(label, row, col)
                col_list.append(label)
            self.result_labels.append(col_list)

    def cleanResultGrid(self):
        """恢复结果区至初始状态"""
        for row in range(16):
            for col in range(16):
                if self.result_labels[row][col] is not None:
                    self.result_labels[row][col].defaultState()


    def startPingTest(self):
        """开始Ping测试"""
        # 验证IP地址输入
        if (self.ui.ipEdit_1.text() != "" and
                self.ui.ipEdit_2.text() != "" and
                self.ui.ipEdit_3.text() != ""):

            self.cleanResultGrid()
            self.ping_finished = False
            self.ui.progressBar.setStyleSheet(self.progressBarStyleSheet)
            self.ui.progressBar.setValue(0)
            self.ui.startPingBtn.setEnabled(False)
            self.ui.stopPingBtn.setEnabled(True)
            self.ip_prefix = (f"{self.ui.ipEdit_1.text()}."
                              f"{self.ui.ipEdit_2.text()}."
                              f"{self.ui.ipEdit_3.text()}.")

            # 创建并启动工作线程
            if self.ping_worker and self.ping_worker.is_running():
                self.ping_worker.stop()

            self.ping_worker = PingWorker(self.ip_prefix)
            self.ping_worker.progress_signal.connect(self.updatePingResult)
            self.ping_worker.finished_signal.connect(self.pingFinished)
            self.ping_worker.ping_error_signal.connect(self.pingError)

            # 启动工作线程
            threading.Thread(target=self.ping_worker.run, daemon=True).start()
            self.ui.progressBar.setFormat("测试中... %p%")
        else:
            w = InfoBar(
                icon=InfoBarIcon.INFORMATION,
                title="提示",
                content="请输入正确的IP地址！",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self.parent
            )
            w.show()

    def stopPingTest(self):
        """停止Ping测试"""
        if self.ping_worker:
            self.ping_worker.stop()

    def updatePingResult(self, idx, result: dict):
        """更新Ping结果"""
        row = idx // 16
        col = idx % 16

        # 获取单元格
        label = self.result_labels[row][col]
        ip = self.ip_prefix + str(idx)

        if result["结果"] == "成功":
            if result["平均延迟"] < 50:
                label.setBackgroundColor(QColor("#50d050"))
            elif result["平均延迟"] < 100:
                label.setBackgroundColor(QColor("#70c8ff"))
            elif result["平均延迟"] < 200:
                label.setBackgroundColor(QColor("#ffff70"))
            else:
                label.setBackgroundColor(QColor("#ff8080"))

            if result["丢包率"] != 0.0:
                label.setBackgroundColor(QColor("#ff8080"))

            label.setToolTip(f"IP地址: {ip}\n测试结果: 成功 | 丢包率: {result["丢包率"]}% | 平均延迟: {result["平均延迟"]}ms")
        else:
            label.setBackgroundColor(QColor("#ff5050"))
            label.setToolTip(f"IP地址: {ip}\n测试结果: 失败 | 目标不可达")

        self.ui.progressBar.setValue(result["进度"] / 254 * 100)

        # 以下是Excel表格数据更新
        self.cell_format(self.ws.cell(row=idx + 1, column=1, value=ip))
        self.cell_format(self.ws.cell(row=idx + 1, column=3, value=result["丢包率"]))
        self.cell_format(self.ws.cell(row=idx + 1, column=4, value=result["最大延迟"]))
        self.cell_format(self.ws.cell(row=idx + 1, column=5, value=result["最小延迟"]))
        self.cell_format(self.ws.cell(row=idx + 1, column=6, value=result["平均延迟"]))

        if result["结果"] == "成功":
            self.cell_format(self.ws.cell(row=idx + 1, column=2, value=result["结果"]), "50d050")
        else:
            self.cell_format(self.ws.cell(row=idx + 1, column=2, value=result["结果"]), "ff5050")

    def pingError(self):
        w = InfoBar(
            icon=InfoBarIcon.ERROR,
            title="错误",
            content="无效IP，向一个无法连接的网络尝试了一个套接字操作！",
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=5000,
            parent=self.parent
        )
        w.show()

    @staticmethod
    def cell_format(cell, fill_color: str = None):
        """统一设置单元格格式"""
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                                bottom=Side(style="thin"))
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, fill_type="solid")

    def pingFinished(self, state: bool):
        """Ping测试完成"""
        if state:
            self.ping_finished = True
            self.ui.progressBar.setValue(100)
            self.ui.progressBar.setFormat("测试完成!")
        else:
            self.ui.progressBar.setFormat("测试终止!")
            self.ui.progressBar.setStyleSheet(f"""
                QProgressBar {{
                    border-radius: 4px;
                    text-align: center;
                    }}
                QProgressBar::chunk {{
                    background: 1px solid #ff5050;
                    border-radius: 4px;
                    border: 1px solid rgba(255, 255, 255, 30%);
                    box-shadow: inset 0 1px 2px rgba(0, 0, 0, 0.1);
                    }}""")
        self.ui.startPingBtn.setEnabled(True)
        self.ui.stopPingBtn.setEnabled(False)

        # 清空工作线程引用
        self.ping_worker = None


class ResultLabel(QLabel):
    def __init__(self, parent):
        super().__init__(parent)
        self.setAlignment(Qt.AlignCenter)
        self.defaultState()

    def setBackgroundColor(self, color: QColor):
        """设置背景色"""
        c = color.name()
        self.setStyleSheet(f"""
                            QLabel {{
                                color: black;
                                border-radius: 4px;
                                background-color: { c };
                                font-weight: regular; }}"""
                            )

    def defaultState(self):
        """默认显示状态"""
        self.setObjectName(f"label_{self.text()}")  # 设置唯一对象名
        self.setToolTip(f"IP地址: xxx.xxx.xxx.{self.text()}\n状态: 等待测试")
        self.installEventFilter(ToolTipFilter(self, showDelay=300, position=ToolTipPosition.BOTTOM))
        self.setBackgroundColor(QColor("#c0c0c0"))